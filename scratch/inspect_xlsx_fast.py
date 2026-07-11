import openpyxl

excel_path = "c:/Users/Antonio/OneDrive/Escritorio/fintech-portfolio-optimized/Datos_Adquisicion_21-22_TOTAL.xlsx"

print("Loading workbook in read_only mode...")
wb = openpyxl.load_workbook(excel_path, read_only=True)
print("Sheets found:", wb.sheetnames)

for name in wb.sheetnames:
    print(f"\n--- Sheet: {name} ---")
    ws = wb[name]
    # Print first 3 rows
    count = 0
    for row in ws.iter_rows(values_only=True):
        print(row)
        count += 1
        if count >= 3:
            break
